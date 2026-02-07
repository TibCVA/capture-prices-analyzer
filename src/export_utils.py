import os
import logging
import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

logger = logging.getLogger("capture_prices.export_utils")


def export_to_excel(all_metrics: list[dict],
                    diagnostics: list[dict],
                    slopes: list[dict],
                    filepath: str):
    """
    Exporte les resultats en fichier Excel avec 3 onglets.
    """
    os.makedirs(os.path.dirname(filepath) or '.', exist_ok=True)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Onglet 1 : Metriques
        df_metrics = pd.DataFrame(all_metrics)
        df_metrics.to_excel(writer, sheet_name='Métriques', index=False)

        # Onglet 2 : Diagnostics
        df_diag = pd.DataFrame(diagnostics)
        df_diag.to_excel(writer, sheet_name='Diagnostics', index=False)

        # Onglet 3 : Slopes
        df_slopes = pd.DataFrame(slopes)
        df_slopes.to_excel(writer, sheet_name='Slopes', index=False)

        # Formatage conditionnel basique
        wb = writer.book
        ws = wb['Métriques']

        # Header en gras
        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

    logger.info(f"Export Excel: {filepath}")


def export_to_gsheets(all_metrics: list[dict],
                      diagnostics: list[dict],
                      slopes: list[dict],
                      spreadsheet_name: str,
                      creds_path: str | None = None) -> str | None:
    """
    Exporte vers Google Sheets. Retourne l'URL du Sheet.
    Retourne None si les credentials sont absentes (graceful fail).
    """
    if creds_path is None:
        creds_path = os.environ.get('GOOGLE_SHEETS_CREDS')

    if not creds_path or not os.path.exists(creds_path):
        logger.warning("Google Sheets creds absentes -- export non disponible")
        return None

    try:
        import gspread
        from google.oauth2.service_account import Credentials

        scopes = ['https://www.googleapis.com/auth/spreadsheets',
                   'https://www.googleapis.com/auth/drive']
        creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
        gc = gspread.authorize(creds)

        sh = gc.create(spreadsheet_name)

        # Metriques
        ws_metrics = sh.sheet1
        ws_metrics.update_title('Métriques')
        df_m = pd.DataFrame(all_metrics)
        ws_metrics.update([df_m.columns.tolist()] + df_m.fillna('').values.tolist())

        # Diagnostics
        ws_diag = sh.add_worksheet(title='Diagnostics', rows=100, cols=20)
        df_d = pd.DataFrame(diagnostics)
        ws_diag.update([df_d.columns.tolist()] + df_d.fillna('').values.tolist())

        # Slopes
        ws_slopes = sh.add_worksheet(title='Slopes', rows=50, cols=15)
        df_s = pd.DataFrame(slopes)
        ws_slopes.update([df_s.columns.tolist()] + df_s.fillna('').values.tolist())

        url = sh.url
        logger.info(f"Export Google Sheets: {url}")
        return url

    except Exception as e:
        logger.error(f"Export Google Sheets echoue: {e}")
        return None
